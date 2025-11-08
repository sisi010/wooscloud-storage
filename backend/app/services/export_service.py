"""
Export Service
Handles data export to various formats (JSON, CSV, Excel)
"""

import json
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting data to various formats"""
    
    def __init__(self, db):
        self.db = db
    
    async def export_to_json(
        self,
        user_id: str,
        collection: str,
        query: Optional[Dict] = None,
        fields: Optional[List[str]] = None
    ) -> bytes:
        """
        Export data to JSON format
        
        Args:
            user_id: User ID
            collection: Collection name
            query: Optional MongoDB query filter
            fields: Optional list of fields to include
            
        Returns:
            JSON data as bytes
        """
        
        # Build query
        base_query = {
            "user_id": user_id,
            "collection": collection
        }
        
        if query:
            base_query.update(query)
        
        # Build projection
        projection = {"_id": 0}  # Exclude MongoDB ID
        if fields:
            for field in fields:
                projection[f"data.{field}"] = 1
        
        # Get data
        cursor = self.db.storage_data.find(base_query, projection)
        data_list = await cursor.to_list(length=None)
        
        # Extract data field
        export_data = []
        for doc in data_list:
            if "data" in doc:
                export_data.append(doc["data"])
        
        # Convert to JSON
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        return json_str.encode('utf-8')
    
    async def export_to_csv(
        self,
        user_id: str,
        collection: str,
        query: Optional[Dict] = None,
        fields: Optional[List[str]] = None
    ) -> bytes:
        """
        Export data to CSV format
        
        Args:
            user_id: User ID
            collection: Collection name
            query: Optional MongoDB query filter
            fields: Optional list of fields to include
            
        Returns:
            CSV data as bytes
        """
        
        # Build query
        base_query = {
            "user_id": user_id,
            "collection": collection
        }
        
        if query:
            base_query.update(query)
        
        # Get data
        cursor = self.db.storage_data.find(base_query)
        data_list = await cursor.to_list(length=None)
        
        if not data_list:
            return b""
        
        # Flatten data and determine all fields
        all_fields = set()
        flattened_data = []
        
        for doc in data_list:
            if "data" in doc:
                flat_row = self._flatten_dict(doc["data"])
                all_fields.update(flat_row.keys())
                flattened_data.append(flat_row)
        
        # Filter fields if specified
        if fields:
            all_fields = [f for f in all_fields if f in fields]
        else:
            all_fields = sorted(list(all_fields))
        
        # Create CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=all_fields)
        writer.writeheader()
        
        for row in flattened_data:
            filtered_row = {k: row.get(k, '') for k in all_fields}
            writer.writerow(filtered_row)
        
        csv_str = output.getvalue()
        return csv_str.encode('utf-8-sig')  # UTF-8 with BOM for Excel
    
    async def export_to_excel(
        self,
        user_id: str,
        collection: str,
        query: Optional[Dict] = None,
        fields: Optional[List[str]] = None
    ) -> bytes:
        """
        Export data to Excel (XLSX) format
        
        Args:
            user_id: User ID
            collection: Collection name
            query: Optional MongoDB query filter
            fields: Optional list of fields to include
            
        Returns:
            Excel file as bytes
        """
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ValueError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Build query
        base_query = {
            "user_id": user_id,
            "collection": collection
        }
        
        if query:
            base_query.update(query)
        
        # Get data
        cursor = self.db.storage_data.find(base_query)
        data_list = await cursor.to_list(length=None)
        
        if not data_list:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = collection[:31]  # Excel limit
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Flatten data and determine all fields
        all_fields = set()
        flattened_data = []
        
        for doc in data_list:
            if "data" in doc:
                flat_row = self._flatten_dict(doc["data"])
                all_fields.update(flat_row.keys())
                flattened_data.append(flat_row)
        
        # Filter fields if specified
        if fields:
            all_fields = [f for f in all_fields if f in fields]
        else:
            all_fields = sorted(list(all_fields))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = collection[:31]  # Excel sheet name limit
        
        # Style for header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write header
        for col_num, field_name in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col_num, value=field_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_num, row_data in enumerate(flattened_data, 2):
            for col_num, field_name in enumerate(all_fields, 1):
                value = row_data.get(field_name, '')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '.') -> Dict:
        """
        Flatten nested dictionary
        
        Example:
            {"user": {"name": "John", "age": 30}}
            -> {"user.name": "John", "user.age": 30}
        """
        items = []
        
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert list to string
                items.append((new_key, json.dumps(v, ensure_ascii=False)))
            else:
                items.append((new_key, v))
        
        return dict(items)
    
    async def get_export_stats(self, user_id: str, collection: str) -> Dict[str, Any]:
        """Get statistics about data to be exported"""
        
        query = {
            "user_id": user_id,
            "collection": collection
        }
        
        count = await self.db.storage_data.count_documents(query)
        
        # Get sample to determine fields
        sample = await self.db.storage_data.find_one(query)
        
        fields = []
        if sample and "data" in sample:
            flat_data = self._flatten_dict(sample["data"])
            fields = list(flat_data.keys())
        
        return {
            "count": count,
            "fields": fields
        }